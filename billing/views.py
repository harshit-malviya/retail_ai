from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, redirect
from .forms import SaleForm, SaleItemFormSet
from django.db import transaction
from accounts.decorators import admin_required

def create_sale(request):
    if request.method == 'POST':
        form = SaleForm(request.POST)
        formset = SaleItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                sale = form.save(commit=False)
                sale.total_amount = 0  # Will calculate from items
                sale.save()

                formset.instance = sale
                items = formset.save()

                total = 0
                for item in items:
                    total += item.get_total_price()

                    # Update stock
                    item.product.stock_quantity -= item.quantity
                    item.product.save()

                sale.total_amount = total
                sale.save()
                sale.customer.last_purchase_date = sale.date  # or timezone.now()
                sale.customer.save()
                customer = sale.customer
                customer.last_purchase = sale.date
                customer.visit_count += 1
                customer.save()
                customer1 = sale.customer
                customer1.last_purchase = sale.date
                customer1.total_spent += sale.total_amount
                customer1.save()

                return redirect('billing:sale_detail', sale.id)  # You'll implement this view next

    else:
        form = SaleForm()
        formset = SaleItemFormSet()
    return render(request, 'billing/create_sale.html', {'form': form, 'formset': formset})

from django.shortcuts import render, get_object_or_404
from .models import Sale

def sale_detail(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)
    return render(request, 'billing/sale_detail.html', {'sale': sale})

from django.http import JsonResponse
from products.models import Product

def get_product_info(request):
    product_id = request.GET.get('product_id')
    try:
        product = Product.objects.get(id=product_id)
        data = {
            'price': float(product.selling_price),
            'stock': product.stock_quantity,
        }
        return JsonResponse(data)
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)

from django.shortcuts import render, get_object_or_404
from customers.models import Customer
from .models import Sale

def customer_purchase_history(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    purchases = Sale.objects.filter(customer=customer).order_by('-date')
    return render(request, 'billing/purchase_history.html', {
        'customer': customer,
        'purchases': purchases
    })

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from customers.models import Customer
from .models import Sale

@admin_required
def export_customer_sales_excel(request, customer_id):
    customer = Customer.objects.get(id=customer_id)
    sales = Sale.objects.filter(customer=customer).prefetch_related('items__product').order_by('-date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase History"

    # Header
    headers = ['Invoice ID', 'Date', 'Product', 'Quantity', 'Price', 'Subtotal']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    # Data rows
    row = 2
    for sale in sales:
        invoice_total = 0
        for item in sale.items.all():
            subtotal = float(item.get_total_price())
            invoice_total += subtotal

            ws[f'A{row}'] = sale.id
            ws[f'B{row}'] = sale.date.strftime('%Y-%m-%d %H:%M')
            ws[f'C{row}'] = item.product.name
            ws[f'D{row}'] = item.quantity
            ws[f'E{row}'] = float(item.price)
            ws[f'F{row}'] = subtotal
            row += 1

        # Total row for this invoice
        ws[f'E{row}'] = "Invoice Total"
        ws[f'F{row}'] = invoice_total
        row += 2  # space before next invoice


    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{customer.name.replace(' ', '_')}_Purchase_History.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

# from reportlab.lib.pagesizes import A4
# from reportlab.pdfgen import canvas
# from io import BytesIO
# from django.http import HttpResponse
# from customers.models import Customer
# from .models import Sale

# def export_customer_sales_pdf(request, customer_id):
#     customer = Customer.objects.get(id=customer_id)
#     sales = Sale.objects.filter(customer=customer).prefetch_related('items__product').order_by('-date')

#     buffer = BytesIO()
#     p = canvas.Canvas(buffer, pagesize=A4)
#     width, height = A4
#     y = height - 50

#     # Title
#     p.setFont("Helvetica-Bold", 14)
#     p.drawString(50, y, f"Purchase History for {customer.name}")
#     y -= 30

#     p.setFont("Helvetica", 10)

#     for sale in sales:
#         if y < 80:
#             p.showPage()
#             y = height - 50

#         # Invoice Header
#         p.setFont("Helvetica-Bold", 11)
#         p.drawString(50, y, f"Invoice #{sale.id} - {sale.date.strftime('%Y-%m-%d %H:%M')}")
#         y -= 20

#         # Table headers
#         p.setFont("Helvetica", 9)
#         p.drawString(60, y, "Product")
#         p.drawString(200, y, "Qty")
#         p.drawString(250, y, "Price")
#         p.drawString(320, y, "Subtotal")
#         y -= 15

#         invoice_total = 0
#         for item in sale.items.all():
#             if y < 60:
#                 p.showPage()
#                 y = height - 50

#             subtotal = item.get_total_price()
#             invoice_total += subtotal

#             p.drawString(60, y, item.product.name)
#             p.drawString(200, y, str(item.quantity))
#             p.drawString(250, y, f"{item.price:.2f}")
#             p.drawString(320, y, f"{subtotal:.2f}")
#             y -= 15

#         # Invoice total
#         if y < 60:
#             p.showPage()
#             y = height - 50

#         p.setFont("Helvetica-Bold", 10)
#         p.drawString(250, y, "Total:")
#         p.drawString(320, y, f"{invoice_total:.2f}")
#         y -= 25  # space before next invoice

#     # Finalize PDF
#     p.showPage()
#     p.save()

#     buffer.seek(0)
#     response = HttpResponse(buffer, content_type='application/pdf')
#     filename = f"{customer.name.replace(' ', '_')}_Purchase_History.pdf"
#     response['Content-Disposition'] = f'attachment; filename={filename}'
#     return response

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from io import BytesIO
from django.http import HttpResponse
from customers.models import Customer
from .models import Sale

@admin_required
def export_customer_sales_pdf(request, customer_id):
    customer = Customer.objects.get(id=customer_id)
    sales = Sale.objects.filter(customer=customer).prefetch_related('items__product').order_by('-date')

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 50

    # Title
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, f"Customer name: {customer.name}")
    y -= 20

    # Contact Info
    p.setFont("Helvetica", 10)
    if customer.phone:
        p.drawString(50, y, f"Phone: {customer.phone}")
        y -= 15
    if customer.email:
        p.drawString(50, y, f"Email: {customer.email}")
        y -= 15
    if customer.address:
        p.drawString(50, y, f"Address: {customer.address[:80]}")
        y -= 15

    y -= 10  # space after header

    for sale in sales:
        if y < 80:
            p.showPage()
            y = height - 50

        # Invoice Header
        p.setFont("Helvetica-Bold", 11)
        p.drawString(50, y, f"Invoice #{sale.id} - {sale.date.strftime('%Y-%m-%d %H:%M')}")
        y -= 20

        # Table headers
        p.setFont("Helvetica", 9)
        p.drawString(60, y, "Product")
        p.drawString(200, y, "Qty")
        p.drawString(250, y, "Price")
        p.drawString(320, y, "Subtotal")
        y -= 15

        invoice_total = 0
        for item in sale.items.all():
            if y < 60:
                p.showPage()
                y = height - 50

            subtotal = item.get_total_price()
            invoice_total += subtotal

            p.drawString(60, y, item.product.name)
            p.drawString(200, y, str(item.quantity))
            p.drawString(250, y, f"{item.price:.2f}")
            p.drawString(320, y, f"{subtotal:.2f}")
            y -= 15

        # Invoice total
        if y < 60:
            p.showPage()
            y = height - 50

        p.setFont("Helvetica-Bold", 10)
        p.drawString(250, y, "Total:")
        p.drawString(320, y, f"{invoice_total:.2f}")
        y -= 25  # space before next invoice

    # Finalize PDF
    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"{customer.name.replace(' ', '_')}_Purchase_History.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response
